from flask import Flask, render_template_string
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import pytz

# Flask app initialization
app = Flask(__name__)

# Path to your Excel file
excel_file = "stock_data_output_1.xlsx"  # Excel filename

def get_font_style(cell):
    """Extract the font style (color, capitalization, etc.) from an Excel cell."""
    styles = []

    # Get font color
    if cell.font and cell.font.color and cell.font.color.type == "rgb":
        font_color = f"#{cell.font.color.rgb[2:]}"  # Skip alpha, take RGB part
        styles.append(f"color: {font_color};")  # Apply the color in hex format

    # Get capitalization (block letters)
    if cell.value and str(cell.value).isupper():
        styles.append("text-transform: uppercase;")

    # Add bold or italic if applicable
    if cell.font and cell.font.bold:
        styles.append("font-weight: bold;")
    if cell.font and cell.font.italic:
        styles.append("font-style: italic;")

    # Combine all styles
    return " ".join(styles)

def get_excel_with_colors(sheet_name):
    # Load the workbook
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook[sheet_name]

    # Extract data and styles
    data = []
    styles = []
    for row in sheet.iter_rows(values_only=False):
        data_row = []
        style_row = []
        for cell in row:
            data_row.append(cell.value)

            # Extract font style and color
            font_style = get_font_style(cell)
            style_row.append(font_style)
        data.append(data_row)
        styles.append(style_row)

    # Convert to DataFrame with styles
    df = pd.DataFrame(data)
    return df, styles

def dataframe_to_html_with_styles(df, styles):
    """Convert a DataFrame to HTML with inline styles."""
    html = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    for i, row in df.iterrows():
        html += "<tr>"
        for j, value in enumerate(row):
            style = styles[i][j] if j < len(styles[i]) else ""
            html += f'<td style="{style}">{value}</td>'
        html += "</tr>"
    html += "</table>"
    return html

def get_nasdaq_status():
    """Determine NASDAQ market hours status based on CET time."""
    cet_timezone = pytz.timezone('CET')
    current_time = datetime.now(cet_timezone)
    market_start_time = current_time.replace(hour=15, minute=0, second=0, microsecond=0)  # 3 PM CET

    if current_time >= market_start_time:
        status = '<span style="color: red; font-weight: bold;">NASDAQ Market hours started</span>'
    else:
        status = '<span style="color: green; font-weight: bold;">NASDAQ Market hours not started</span>'

    return status

@app.route('/')
def display_excel():
    # Timezone sheet
    timezone_df, timezone_styles = get_excel_with_colors('Timezonesheet')
    timezone_html = dataframe_to_html_with_styles(timezone_df, timezone_styles)

    # MyStock sheet
    my_stock_df, my_stock_styles = get_excel_with_colors('MyStock')
    my_stock_html = dataframe_to_html_with_styles(my_stock_df, my_stock_styles)

    # NewStock sheet
    new_stock_df, new_stock_styles = get_excel_with_colors('NewStock')
    new_stock_html = dataframe_to_html_with_styles(new_stock_df, new_stock_styles)

    # Watched stock sheet
    watched_stock_df, watched_stock_styles = get_excel_with_colors('WatchedStock')
    watched_stock_html = dataframe_to_html_with_styles(watched_stock_df, watched_stock_styles)

    # Get NASDAQ status
    nasdaq_status = get_nasdaq_status()

    # HTML template with sections
    html_template = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Excel Viewer with Colors</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 0;
                padding: 0;
            }}
            .container {{
                display: flex;
                flex-direction: column;
                align-items: center;
                width: 100%;
                height: 100vh;
            }}
            .section-container {{
                display: flex;
                justify-content: space-between;
                width: 100%;
                height: calc(100% - 100px);
            }}
            .section {{
                flex-basis: 33%;
                padding: 10px;
                border: 1px solid #ddd;
                text-align: center;
                overflow-y: auto; /* Enable vertical scrolling */
                box-sizing: border-box;
                height: 100%; /* Ensure it occupies the full height */
            }}
            .section h3 {{
                margin-top: 0;
            }}
            table {{
                width: 100%;
                table-layout: fixed;
                word-wrap: break-word;
            }}
            th, td {{
                padding: 8px;
                text-align: left;
                border: 1px solid #ddd;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <!-- Timezone Section -->
            <div>{timezone_html}</div>

            <!-- NASDAQ Status Section -->
            <div style="text-align: center; margin: 20px;">
                <h2>{nasdaq_status}</h2>
            </div>

            <!-- Main Content Section -->
            <div class="section-container">
                <!-- NewStock Section -->
                <div class="section">
                    <h3>NewStock</h3>
                    {new_stock_html}
                </div>

                <!-- Watched Stock Section -->
                <div class="section">
                    <h3>Watched Stock</h3>
                    {watched_stock_html}
                </div>

                <!-- MyStock Section -->
                <div class="section">
                    <h3>MyStock</h3>
                    {my_stock_html}
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    return render_template_string(html_template)

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5000, debug=True)
